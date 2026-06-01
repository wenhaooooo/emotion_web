import asyncio
import base64
import io
import json
import logging
import tempfile
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import cv2
import numpy as np
import soundfile as sf
from fastapi import APIRouter, UploadFile, File, HTTPException, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse, FileResponse

from api.schemas import JobStatus, JobResult
from core.face_recognizer import get_face_recognizer
from tasks.manager import JobManager, run_processing_job

logger = logging.getLogger(__name__)

router = APIRouter()

BASE_DIR = Path(__file__).parent.parent
UPLOADS_DIR = BASE_DIR / "storage" / "uploads"
JOBS_DIR = BASE_DIR / "storage" / "jobs"

job_manager = JobManager(JOBS_DIR)

# 线程池用于模型推理
thread_pool = ThreadPoolExecutor(max_workers=4)

# 实时识别音频配置
AUDIO_WINDOW_SECONDS = 3      # 音频窗口时长
AUDIO_SLIDE_SECONDS = 1.5     # 滑动步长
AUDIO_SAMPLE_RATE = 16000     # 采样率
SILENCE_RMS_THRESHOLD = 0.03  # 静音检测阈值（RMS 能量），低于此值视为安静

ALLOWED_EXTENSIONS = {".mp4", ".avi", ".mov", ".mkv", ".flv", ".wmv"}


@router.post("/upload", response_model=JobStatus)
async def upload_video(file: UploadFile = File(...)):
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Unsupported file format: {ext}. Allowed: {', '.join(ALLOWED_EXTENSIONS)}")

    # Save uploaded file
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    save_path = UPLOADS_DIR / file.filename
    content = await file.read()
    save_path.write_bytes(content)

    # Create job
    job_id = job_manager.create_job(str(save_path), file.filename)

    # Start background processing
    thread = threading.Thread(
        target=run_processing_job,
        args=(job_manager, job_id, str(save_path), file.filename, JOBS_DIR),
        daemon=True,
    )
    thread.start()

    return job_manager.get_status(job_id)


@router.get("/jobs/{job_id}/progress")
async def job_progress(job_id: str, request: Request):
    status = job_manager.get_status(job_id)
    if status is None:
        raise HTTPException(404, "Job not found")

    if status.status in ("done", "failed"):
        async def single_event():
            yield f"data: {json.dumps(status.model_dump())}\n\n"
        return StreamingResponse(single_event(), media_type="text/event-stream")

    q = job_manager.subscribe_sse(job_id)

    async def event_generator():
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    data = await asyncio.to_thread(q.get, timeout=30)
                    yield f"data: {data}\n\n"
                    parsed = json.loads(data)
                    if parsed.get("status") in ("done", "failed"):
                        break
                except Exception:
                    yield ": keepalive\n\n"
        finally:
            job_manager.cleanup_sse(job_id, q)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@router.get("/jobs/{job_id}/result", response_model=JobResult)
async def job_result(job_id: str):
    result = job_manager.get_result(job_id)
    if result is None:
        status = job_manager.get_status(job_id)
        if status is None:
            raise HTTPException(404, "Job not found")
        if status.status == "failed":
            raise HTTPException(500, f"Job failed: {status.message}")
        raise HTTPException(202, "Job still processing")
    return result


@router.get("/jobs/{job_id}/original-video")
async def job_original_video(job_id: str):
    video_path = job_manager.get_video_path(job_id)
    if video_path is None or not Path(video_path).is_file():
        raise HTTPException(404, "Original video not found")
    return FileResponse(video_path, media_type="video/mp4")


@router.get("/jobs/{job_id}/video/{segment_index}")
async def job_video_segment(job_id: str, segment_index: int):
    segments_dir = JOBS_DIR / job_id / "segments"
    if not segments_dir.exists():
        raise HTTPException(404, "Segments not found")

    matches = list(segments_dir.glob(f"*_seg{segment_index:04d}.mp4"))
    if not matches:
        raise HTTPException(404, f"Segment {segment_index} not found")

    return FileResponse(matches[0], media_type="video/mp4")


def _format_time(seconds: float) -> str:
    """将秒数格式化为 HH:MM:SS 格式"""
    total_seconds = int(seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secs = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


@router.get("/jobs/{job_id}/export")
async def job_export_excel(job_id: str):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    result = job_manager.get_result(job_id)
    if result is None:
        status = job_manager.get_status(job_id)
        if status is None:
            raise HTTPException(404, "Job not found")
        if status.status == "failed":
            raise HTTPException(500, f"Job failed: {status.message}")
        raise HTTPException(202, "Job still processing")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "情感识别结果"

    # Header
    headers = ["#", "开始时间", "结束时间", "时长", "文本",
               "emotion2vec标签", "emotion2vec标签名", "教师标签", "教师标签名", "置信度"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for seg in result.segments:
        row = [
            seg.index + 1,
            _format_time(seg.start_time),
            _format_time(seg.end_time),
            _format_time(seg.duration),
            seg.text,
            seg.emotion2vec_label,
            seg.emotion2vec_label_name,
            seg.label,
            seg.label_name_cn,
            round(seg.confidence, 4),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=seg.index + 2, column=col, value=val)
            cell.border = thin_border

    # Column widths
    widths = [5, 10, 10, 10, 40, 12, 14, 8, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A1:J{len(result.segments) + 1}"

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{result.video_name}_emotions.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ==================== 实时情感识别 WebSocket ====================


def _process_video_frame(base64_data: str) -> dict:
    """处理视频帧：base64 解码 → 人脸检测 → 表情识别"""
    face_recognizer = get_face_recognizer()
    if not face_recognizer.is_loaded:
        return {"type": "face_emotion", "face_detected": False}

    try:
        # base64 解码为图像
        img_bytes = base64.b64decode(base64_data)
        nparr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        if frame is None:
            return {"type": "face_emotion", "face_detected": False}

        # 推理
        result = face_recognizer.predict(frame)
        result["type"] = "face_emotion"
        return result
    except Exception as e:
        logger.error(f"视频帧处理失败: {e}")
        return {"type": "face_emotion", "face_detected": False}


def _process_audio_buffer(audio_buffer: np.ndarray) -> dict:
    """处理音频缓冲：PCM → WAV → emotion2vec 推理"""
    from core.recognizer import _get_predictor
    from core.face_recognizer import map_emotion2vec_to_7class

    try:
        # 静音检测：计算 RMS 能量
        rms = float(np.sqrt(np.mean(audio_buffer ** 2)))
        if rms < SILENCE_RMS_THRESHOLD:
            return {"type": "voice_emotion", "silence": True}

        # 保存为临时 WAV 文件
        with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as tmp:
            tmp_path = tmp.name
            sf.write(tmp_path, audio_buffer, AUDIO_SAMPLE_RATE)

        # emotion2vec 推理
        predictor = _get_predictor()
        raw_result = predictor.predict(tmp_path)

        # 清理临时文件
        import os
        try:
            os.remove(tmp_path)
        except OSError:
            pass

        # 映射到 7 类
        result = map_emotion2vec_to_7class(raw_result)
        result["type"] = "voice_emotion"
        return result
    except Exception as e:
        logger.error(f"音频处理失败: {e}")
        return {"type": "voice_emotion", "silence": True}


@router.websocket("/ws/realtime")
async def websocket_realtime(websocket: WebSocket):
    """实时情感识别 WebSocket 端点"""
    await websocket.accept()

    # 创建实时识别 Job
    job_id = uuid.uuid4().hex[:12]
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    await websocket.send_json({"type": "connected", "job_id": job_id})
    logger.info(f"实时识别 WebSocket 已连接, job_id={job_id}")

    # 音频缓冲区
    audio_buffer = np.array([], dtype=np.float32)

    # 结果累积
    results = []
    start_time = time.time()

    # 跟踪并发任务
    pending_tasks = set()

    try:
        while True:
            raw = await websocket.receive_text()
            message = json.loads(raw)
            msg_type = message.get("type")

            if msg_type == "video_frame":
                # 在线程池中处理视频帧
                loop = asyncio.get_event_loop()
                task = loop.run_in_executor(
                    thread_pool,
                    _process_video_frame,
                    message["data"],
                )
                pending_tasks.add(task)
                task.add_done_callback(pending_tasks.discard)

                # 异步等待结果并推送
                async def send_face_result(t):
                    try:
                        result = await t
                        results.append({
                            "timestamp": time.time() - start_time,
                            "source": "face",
                            **result,
                        })
                        await websocket.send_json(result)
                    except Exception as e:
                        logger.error(f"发送面部结果失败: {e}")

                asyncio.create_task(send_face_result(task))

            elif msg_type == "audio_chunk":
                # base64 解码音频数据（PCM float32）
                try:
                    audio_bytes = base64.b64decode(message["data"])
                    pcm_data = np.frombuffer(audio_bytes, dtype=np.float32)
                    audio_buffer = np.concatenate([audio_buffer, pcm_data])
                except Exception as e:
                    logger.error(f"音频解码失败: {e}")
                    continue

                # 当缓冲区达到窗口大小时触发推理
                window_samples = AUDIO_WINDOW_SECONDS * AUDIO_SAMPLE_RATE
                if len(audio_buffer) >= window_samples:
                    # 取出一个窗口
                    window = audio_buffer[:window_samples].copy()

                    # 滑动窗口：保留后半部分
                    slide_samples = int(AUDIO_SLIDE_SECONDS * AUDIO_SAMPLE_RATE)
                    audio_buffer = audio_buffer[slide_samples:]

                    # 在线程池中处理音频
                    loop = asyncio.get_event_loop()
                    task = loop.run_in_executor(
                        thread_pool,
                        _process_audio_buffer,
                        window,
                    )
                    pending_tasks.add(task)
                    task.add_done_callback(pending_tasks.discard)

                    async def send_voice_result(t):
                        try:
                            result = await t
                            results.append({
                                "timestamp": time.time() - start_time,
                                "source": "voice",
                                **result,
                            })
                            await websocket.send_json(result)
                        except Exception as e:
                            logger.error(f"发送语音结果失败: {e}")

                    asyncio.create_task(send_voice_result(task))

    except WebSocketDisconnect:
        logger.info(f"实时识别 WebSocket 断开, job_id={job_id}")
    except Exception as e:
        logger.error(f"WebSocket 错误: {e}")
    finally:
        # 等待所有待处理任务完成
        if pending_tasks:
            await asyncio.gather(*pending_tasks, return_exceptions=True)

        # 保存结果
        duration = time.time() - start_time
        realtime_result = {
            "job_id": job_id,
            "mode": "realtime",
            "duration_seconds": round(duration, 2),
            "start_time": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime(start_time)),
            "end_time": time.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "timeline": [
                {
                    "timestamp": r.get("timestamp", 0),
                    "source": r.get("source", "face"),
                    "emotion": r.get("emotion", "neutral"),
                    "confidence": r.get("confidence", 0),
                    "probabilities": r.get("probabilities", {}),
                }
                for r in results
            ],
            "summary": _compute_summary(results),
        }

        result_path = job_dir / "result.json"
        result_path.write_text(json.dumps(realtime_result, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info(f"实时识别结果已保存: {result_path}")


def _compute_summary(results: list) -> dict:
    """计算情感分布统计"""
    emotions = ["angry", "disgust", "fear", "happy", "neutral", "sad", "surprise"]
    summary = {
        "face": {e: 0 for e in emotions},
        "voice": {e: 0 for e in emotions},
    }
    for r in results:
        source = r.get("source", "face")
        emotion = r.get("emotion", "neutral")
        if source in summary and emotion in summary[source]:
            summary[source][emotion] += 1
    return summary
